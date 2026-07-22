#!/usr/bin/env python3
"""Import CCPL teams/players from CCPL.xlsx into src/data/teams.json."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path.home() / "Downloads" / "CCPL.xlsx"
DEFAULT_OUT = ROOT / "src" / "data" / "teams.json"
COUNT_SUFFIX = re.compile(r"\((\d+)\)\s*$")
CAPTAIN_SUFFIX = re.compile(r"\s*\(C\)\s*$", re.I)


def normalize_team_name(raw: str) -> str:
    name = raw.strip()
    name = re.sub(r"^Cricket team name\s*[—–-]\s*", "", name, flags=re.I)
    name = re.sub(COUNT_SUFFIX, "", name).strip()
    if name.lower() in ("lifecycle cricket team", "lifecycle cricket team"):
        return "Lifecycle Cricket Team"
    return name


def parse_team_header(cell) -> tuple[str, int] | None:
    if not cell:
        return None
    text = str(cell).strip()
    count_match = COUNT_SUFFIX.search(text)
    if not count_match:
        return None
    expected = int(count_match.group(1))
    name = normalize_team_name(text)
    if not name:
        return None
    return name, expected


def load_existing_captains(out_path: Path) -> dict[str, dict]:
    if not out_path.exists():
        return {}
    try:
        teams = json.loads(out_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    return {
        t["name"]: {
            "captain": t.get("captain", ""),
            "captainEmail": t.get("captainEmail", ""),
            "status": t.get("status", ""),
            "need": t.get("need", ""),
        }
        for t in teams
        if isinstance(t, dict) and t.get("name")
    }


def parse_team_details_sheet(wb) -> dict[str, dict]:
    if "Team Details" not in wb.sheetnames:
        return {}
    details: dict[str, dict] = {}
    for row in wb["Team Details"].iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        name = normalize_team_name(str(row[1]))
        details[name] = {
            "captain": str(row[2]).strip() if row[2] else "",
            "captainEmail": str(row[3]).strip() if row[3] else "",
            "status": str(row[5]).strip() if len(row) > 5 and row[5] else "",
            "need": str(row[6]).strip() if len(row) > 6 and row[6] else "",
        }
    return details


def parse_workbook(xlsx_path: Path, out_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    team_details = parse_team_details_sheet(wb)
    existing = load_existing_captains(out_path)

    sheet_name = "CCPL" if "CCPL" in wb.sheetnames else wb.sheetnames[0]
    all_rows = list(wb[sheet_name].iter_rows(values_only=True))

    headers: list[dict] = []
    for ri, row in enumerate(all_rows):
        for ci, cell in enumerate(row):
            parsed = parse_team_header(cell)
            if parsed:
                name, expected = parsed
                headers.append(
                    {
                        "row": ri,
                        "col": ci,
                        "name": name,
                        "expectedCount": expected,
                    }
                )

    headers.sort(key=lambda h: (h["row"], h["col"]))
    teams: list[dict] = []

    for i, header in enumerate(headers):
        name_col = header["col"]
        email_col = header["col"] + 1
        start_row = header["row"] + 3
        end_row = len(all_rows)

        for nh in headers[i + 1 :]:
            if nh["row"] > start_row:
                end_row = min(end_row, nh["row"])
                break

        players: list[dict] = []
        seen: set[str] = set()
        captain_from_marker = ""

        for ri in range(start_row, end_row):
            row = all_rows[ri]
            if name_col >= len(row) or not row[name_col]:
                continue
            player_name = str(row[name_col]).strip()
            if not player_name or player_name.lower() in ("player", "email / name"):
                continue
            if parse_team_header(player_name):
                break

            if CAPTAIN_SUFFIX.search(player_name):
                captain_from_marker = CAPTAIN_SUFFIX.sub("", player_name).strip()
                player_name = captain_from_marker

            email = None
            if email_col < len(row) and row[email_col]:
                email_raw = str(row[email_col]).replace("\xa0", " ").strip()
                if "@" in email_raw:
                    email = email_raw

            key = player_name.lower()
            if key in seen:
                continue
            seen.add(key)
            players.append({"name": player_name, "email": email})

        meta = (
            team_details.get(header["name"])
            or existing.get(header["name"])
            or {}
        )
        captain = captain_from_marker or meta.get("captain", "")
        captain_email = meta.get("captainEmail", "")

        if not captain and captain_email:
            captain = next(
                (p["name"] for p in players if p.get("email") == captain_email),
                "",
            )
        if not captain and players:
            captain = players[0]["name"]

        words = [w for w in header["name"].split() if w and w.lower() not in ("the", "xi")]
        short_name = "".join(w[0] for w in words)[:3].upper() or header["name"][:3].upper()

        teams.append(
            {
                "name": header["name"],
                "shortName": short_name,
                "captain": captain,
                "captainEmail": captain_email,
                "status": meta.get("status", "Confirmed"),
                "need": meta.get("need", "OK"),
                "players": players,
            }
        )

    wb.close()
    return teams


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUT

    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    teams = parse_workbook(xlsx_path, out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(teams, indent=2), encoding="utf-8")

    player_total = sum(len(t["players"]) for t in teams)
    print(f"Imported {len(teams)} teams, {player_total} players -> {out_path}")
    for t in teams:
        print(f"  {t['name']}: {len(t['players'])} players, captain={t['captain']}")


if __name__ == "__main__":
    main()
